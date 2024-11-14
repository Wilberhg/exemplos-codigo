from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from src.db.action import cria_registro, cria_controle, consulta_fila, atualiza_status_controle

def carga_banco():
    wb = load_workbook("./input/challenge.xlsx")
    ws = wb.active
    cabecalho = ('first_name', 'last_name', 'company_name', 'role_company', 'address', 'email', 'phone_number')
    for row in ws.iter_rows(min_row=2, max_row=11, max_col=7, values_only=True):
        campos_registro = dict(zip(cabecalho, row))
        id = cria_registro(campos_registro)
        cria_controle({'id_controle_analise': id})
    wb.close()
        
def executa_challenge():
    dados_cadastrais = consulta_fila()
    if dados_cadastrais:
        driver = webdriver.Chrome()
        driver.maximize_window()
        driver.implicitly_wait(30)
        driver.get("https://rpachallenge.com/")
        for dados in dados_cadastrais:
            driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelFirstName"]').send_keys(dados.AnaliseCadastro.first_name)
            driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelRole"]').send_keys(dados.AnaliseCadastro.role_company)
            driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelAddress"]').send_keys(dados.AnaliseCadastro.address)
            driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelCompanyName"]').send_keys(dados.AnaliseCadastro.company_name)
            driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelPhone"]').send_keys(dados.AnaliseCadastro.phone_number)
            driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelLastName"]').send_keys(dados.AnaliseCadastro.last_name)
            driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelEmail"]').send_keys(dados.AnaliseCadastro.email)
            driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]').click()
            atualiza_status_controle(dados.ControleAnaliseCadastro.id, {'dado_cadastrado': True})
    

if __name__ == "__main__":
    carga_banco()
    executa_challenge()